import os
from datetime import datetime as dt, timedelta
from time import sleep

import json
from openpyxl import styles, Workbook
import sqlite3
import telebot
from telebot import types

# Имена файлов, использующихся в программе
DB_NAME = "Table.sqlite"
EXCEL_TABLE = "temp_table.xlsx"
USERS_NAME = 'users.json'
COMPANIES_NAME = 'companies.json'

# Константы, по совместительству заголовки в таблице Excel и БД, ключи в файлах json
# Лучше не менять, иначе файл users.json и БД нужно будет перезаписать
HEADERS = ['Company', 'Address', 'Username', 'User_id', 'Phone', 'Counter', 'Data', 'Datetime']
COMPANY, ADDRESS, USERNAME, USER_ID, PHONE, COUNTER, DATA, DATETIME = HEADERS

INJECTIONS = r"""'or 1=1;' or 1=1--;' or 1=1#;' or 1=1/*;' --;' #;'/*;' or '1'='1;' or '1'='1'--;' or '1'='1'#;
' or '1'='1'/*;'or 1=1 or ''=';' or 1=1;' or 1=1--;' or 1=1#;' or 1=1/*;') or ('1'='1;') or ('1'='1'--;
') or ('1'='1'#;') or ('1'='1'/*;') or '1'='1;') or '1'='1'--;') or '1'='1'#;') or '1'='1'/*;" --;" #;
"/*;" or "1"="1;" or "1"="1"--;" or "1"="1"#;" or "1"="1"/*;"or 1=1 or ""=";" or 1=1;" or 1=1--;" or 1=1#;
" or 1=1/*;") or ("1"="1;") or ("1"="1"--;") or ("1"="1"#;") or ("1"="1"/*;") or "1"="1;") or "1"="1"--;
") or "1"="1"#;") or "1"="1"/*""".replace('\n', '').split(';')

POSITIVE_ANSWERS = ['yes', 'y', 'да', 'д', '1', 'дп', 'lf']  # Ответы, которые мы принимаем за положительный ответ
RUS = 'абвгдеёжзийклмнопрстуфхцчшщъыьэюя'
ENG = 'abcdefghijklmnopqrstuvwxyz'
DIGITS = '1234567890'
PUNCTUATION = ' -.,()/+_\n'
ALLOWED_SIMBOLS = DIGITS + ENG + ENG.upper() + RUS + RUS.upper() + PUNCTUATION


def dump(obj, filename):
    """Функция для простого внесения данных в файлы"""
    json.dump(obj, open(filename, 'w', encoding='UTF-8'), ensure_ascii=False, indent=4)


#          Sophia,     Maksim
ADMINS = [979923466, 1089524173]
TOKEN = '5253767532:AAF9DZ2obpuMVKiQD_VHmskA5WTtkjkys3k'

bot = telebot.TeleBot(TOKEN)

# Открытие БД
conn = sqlite3.connect(DB_NAME, check_same_thread=False)

# Создание и открытие json файла со списком зарегистрированных пользователей
if not os.path.exists(USERS_NAME):
    dump({}, USERS_NAME)
users = json.load(open(USERS_NAME, 'r', encoding='utf-8'))

# Создание и открытие файла со списком компаний
if not os.path.exists(COMPANIES_NAME):
    dump({}, COMPANIES_NAME)
companies = json.load(open(COMPANIES_NAME, 'r', encoding='UTF-8'))

# Словарь, в котором будет содержаться информация, которую вносят пользователи в данный момент
recording_data = {}


def get_date() -> str:
    """Получение строки с датой и временем сейчас"""
    return dt.now().strftime("%Y-%m-%d %H:%M:%S")


def get_changes(old: dict, new: dict = None) -> str:
    """Получение строки с изменениями в словаре. Проверяет только значения, существующие в обоих словарях"""

    data = [(key, old[key], (new or old)[key]) for key in old if not new or key in new]
    return "\n".join([f"{i[0]}: {i[1] + ' => ' + i[2] if i[1] != i[2] else i[1]}" for i in data])


def check_number(number) -> str:
    """Изменение номера телефона по формату"""
    number = ''.join(number.replace('(', '').replace(')', '').replace('-', '').split())
    if number[0] == '8':
        number = '+7' + number[1:]
    assert len(number) == 12
    assert number[:2] == '+7'
    assert number[2:].isdigit()
    return number


def check_data(data) -> bool:
    """Проверка, соответствуют ли введенные данные ПУ допустимому формату"""
    try:
        float(data), int(data.split('.')[0]), int(data.split('.')[-1])
        return True
    except ValueError:
        return False


def log(message, symbols=ALLOWED_SIMBOLS, start_call=False) -> bool:
    """Вывод в консоль уведомления о сообщении боту + Проверка сообщения (выход, атака)"""
    name = ((message.from_user.last_name or ' ') + ' ' + message.from_user.first_name).strip()
    if str(message.from_user.id) not in users:
        name += f' (id {message.from_user.id})'
    print(f'{get_date()} - {name}: "{message.text}"')
    # print(recording_data)

    if not message.text:
        error_text = 'Текст сообщения не должен быть пустым'
    elif message.text == '/exit':
        error_text = 'Выход в меню'
    elif '/' in message.text and not start_call:
        start(message)
        return True

    elif any(i in message.text for i in INJECTIONS):
        error_text = 'Сообщение содержит недопустимые символы'
        for admin in ADMINS:
            bot.send_message(admin, f'Сообщение "{message.text}" от пользователя '
                                    f'id{message.from_user.id} показалось подозрительным')

    elif any(i not in symbols for i in message.text):
        error_text = 'Сообщение содержит недопустимые символы'
    elif len(message.text) > 255:
        error_text = 'Сообщение не должно превышать длину 255 символов'
    else:
        return False

    bot.send_message(message.from_user.id, error_text)
    print_commands(message)
    return True


def make_bool_keyboard(one_time=True):
    """Возвращает клавиатуру, состоящую из кнопок "Да" и "Нет\""""
    keyboard = types.ReplyKeyboardMarkup(True, one_time)
    keyboard.add(types.KeyboardButton('Да'), types.KeyboardButton('Нет'))
    return keyboard


def make_keyboard(values, one_time=True):
    """Возвращает клавиатуру, содержащую кнопки со значениями values"""
    keyboard = types.ReplyKeyboardMarkup(True, one_time)
    for value in values:
        key1 = types.KeyboardButton(value)
        keyboard.add(key1)
    return keyboard


def print_commands(message):
    text = f'''
Воспользуйтесь функциями меню:
/create_entry - внесение показания прибора учета
/get_entries - получение записанных показаний по приборам учета
/add_counter - регистрация прибора учета по вашему адресу
/remove_counter - удаление прибора учета по вашему адресу
/edit_user - редактирование вашего профиля
/message_to_admin - отправление сообщения администратору
/exit - завершения работы
'''
    if message.from_user.id in ADMINS:
        text += '''
/get_records - получение показаний за последние 30 дней в виде Excel таблицы
/add_company - добавление компании
/get_companies - просмотр всех зарегистрированных компаний
/remove_user - удаление зарегистрированного пользователя по id
/edit_user_by_id - редактирование данных зарегистрированного пользователя по id
/message_to_user - отправление сообщения пользователю по id
'''
    bot.send_message(message.from_user.id, text)


@bot.message_handler(content_types=['text'])
def start(message):
    """Изначальная функция, принимающая запросы пользователя"""
    if log(message, start_call=True):
        return

    user_id = message.from_user.id

    if message.text == '/message_to_admin':
        bot.send_message(user_id, 'Введите текст сообщения')
        bot.register_next_step_handler(message, message_to_admin)

    elif str(user_id) not in users:
        if message.text == '/edit_user':
            message.text = 'Да'
            if_registration(message)

        else:
            bot.send_message(user_id, "Вы не зарегистрированы. Хотите зарегистрироваться?",
                             reply_markup=make_bool_keyboard())
            bot.register_next_step_handler(message, if_registration)

    elif message.text == '/create_entry':
        create_entry(message)

    elif message.text == '/edit_user':
        # Подтверждаем регистрацию
        bot.send_message(user_id, 'Вы уже зарегистрированы. Ваши данные будут заменены. Вы уверены?',
                         reply_markup=make_bool_keyboard())
        bot.register_next_step_handler(message, if_registration)

    elif message.text == '/get_entries':
        get_entries(message)

    elif message.text == '/add_counter':
        bot.send_message(user_id, 'Введите номер регистрируемого прибора учёта')
        bot.register_next_step_handler(message, add_counter)

    elif message.text == '/remove_counter':
        user = users[str(user_id)]
        counters = companies[user[COMPANY]][user[ADDRESS]]
        bot.send_message(user_id, 'Выберите удаляемый прибор учёта',
                         reply_markup=make_keyboard(counters))
        bot.register_next_step_handler(message, remove_counter)

    elif user_id in ADMINS and message.text == '/add_company':
        bot.send_message(user_id, 'Введите название регистрируемой компании')
        bot.register_next_step_handler(message, add_company)

    elif user_id in ADMINS and message.text == '/get_companies':
        get_companies(message)

    elif user_id in ADMINS and message.text == '/remove_user':
        bot.send_message(user_id, 'Введите id пользователя')
        bot.register_next_step_handler(message, remove_user_by_id)

    elif user_id in ADMINS and message.text == '/edit_user_by_id':
        bot.send_message(user_id, 'Введите id пользователя')
        bot.register_next_step_handler(message, edit_user_by_id)

    elif user_id in ADMINS and message.text == '/message_to_user':
        bot.send_message(user_id, 'Введите id получателя')
        bot.register_next_step_handler(message, message_to_user)

    elif user_id in ADMINS and message.text == '/get_records':
        get_records(message)

    # Обработка сообщений, не содержащих команд
    else:
        print_commands(message)


def message_to_admin(message):
    if log(message):
        return

    text = f'Пользователь {message.from_user.first_name} (id{message.from_user.id}) ' \
           f'отправил администраторам сообщение: "{message.text}"'
    for admin_id in ADMINS:
        bot.send_message(admin_id, text)

    bot.send_message(message.from_user.id, 'Сообщение отправлено')


def message_to_user(message):
    if log(message):
        return

    if message.text.lower() == 'self':
        message.text = message.from_user.id

    if not message.text.isdigit():
        bot.send_message(message.from_user.id, 'Неправильный id')
        print_commands(message)
        return

    recording_data[message.from_user.id] = int(message.text)

    bot.send_message(message.from_user.id, 'Введите сообщение')
    bot.register_next_step_handler(message, message_to_user_text)


def message_to_user_text(message):
    if log(message):
        del recording_data[message.from_user.id]
        return

    user_id = message.from_user.id

    try:
        bot.send_message(recording_data[user_id], f'Администратор отправил вам сообщение: "{message.text}"')
    except Exception as error:
        print(error)
        if 'chat not found' in str(error):
            bot.send_message(user_id, 'Чата с этим пользователем не существует '
                                      '(пользователь ни разу не писал боту/неверный id)')
        elif 'bot was blocked by the user' in str(error):
            bot.send_message(user_id, 'Пользователь запретил боту отправлять ему сообщения')
        else:
            bot.send_message(user_id, str(error))

    else:
        bot.send_message(user_id, 'Сообщение отправлено')


def create_entry(message):
    user_id = message.from_user.id
    user = users[str(user_id)]

    # Предлагаем пользователю список счётчиков по этому адресу
    counters = companies[user[COMPANY]][user[ADDRESS]]

    # Если счетчик всего один, не спрашиваем пользователя
    if len(counters) == 0:
        bot.send_message(user_id, 'Нет зарегистрированных приборов учёта. '
                                  'Для регистрации введите /add_counter')
        return

    if len(counters) == 1:
        message.text = list(counters)[0]
        get_counter(message)
        return

    bot.send_message(user_id, "Выберите номер прибора учёта из списка",
                     reply_markup=make_keyboard(counters))
    bot.register_next_step_handler(message, get_counter)


def get_records(message):
    date_from = dt.now() - timedelta(days=30)
    date_to = dt.now()
    cursor = conn.cursor()
    request = (f"SELECT * FROM records WHERE {DATETIME} BETWEEN '{date_from.strftime('%Y-%m-%d')}'"
               f" AND '{get_date()}' ORDER BY {DATETIME}")
    result = cursor.execute(request).fetchall()
    cursor.close()

    workbook = Workbook()
    sheet = workbook.worksheets[0]
    for j, header in enumerate(HEADERS, 1):
        cell = sheet.cell(1, j)
        cell.value = header.capitalize()
        cell.font = styles.Font(bold=True)
        cell.alignment = styles.Alignment(horizontal='center')

    for i, record in enumerate(result, 2):
        for j, value in enumerate(record[1:], 1):
            cell = sheet.cell(i, j)
            cell.value = value

    workbook.save(EXCEL_TABLE)
    workbook.close()

    filename = f"Records for {date_from.strftime('%d.%m.%Y')} - {date_to.strftime('%d.%m.%Y')}"
    bot.send_document(message.chat.id, open(EXCEL_TABLE, 'rb').read(),
                      visible_file_name=filename + '.xlsx')
    os.remove(EXCEL_TABLE)


def get_companies(message):
    """Администратору выводится список всех компаний"""
    if companies:
        text = 'Список всех зарегистрированных компаний:\n'
        text += '\n'.join(companies)
    else:
        text = 'Нет зарегистрированных компаний'
    bot.send_message(message.from_user.id, text if len(text) < 4096 else text[:4093] + '...')


def get_entries(message):
    """Пользователю выводится список всех счётчиков и их текущих показаний"""
    user = users[str(message.from_user.id)]
    if companies[user[COMPANY]][user[ADDRESS]]:
        text = 'Последние полученные показания по вашим ПУ:\n'
        for i in companies[user[COMPANY]][user[ADDRESS]].items():
            text += f'"{i[0]}": {i[1] or "Нет показаний"}\n'
    else:
        text = 'Нет зарегистрированных приборов учета'
    bot.send_message(message.from_user.id, text if len(text) < 4096 else text[:4093] + '...')


def if_registration(message):
    """Проверка, точно ли пользователь хочет зарегистрироваться
    Регистрация происходит в несколько этапов.
    Компания => Номер телефона => Имя
    Эти данные впоследствии будут заноситься в таблицу, когда этот пользователь вводит показания"""
    if log(message):
        return

    if message.text.lower() in POSITIVE_ANSWERS:
        message.text = 'self'
        edit_user_by_id(message)

    else:
        if str(message.from_user.id) not in users:
            bot.send_message(message.from_user.id, 'Регистрация отменена')
        else:
            bot.send_message(message.from_user.id, 'Выход в меню')
            print_commands(message)


def edit_user_by_id(message):
    """Администратор редактирует данные зарегистрированного пользователя по его id"""
    if log(message):
        return

    user_id = message.from_user.id
    cur_user_id = message.text

    if cur_user_id.lower() == 'self':
        cur_user_id = str(user_id)

    if cur_user_id not in users and str(user_id) != cur_user_id:
        bot.send_message(user_id, 'Пользователь с этим id не зарегистрирован')
        print_commands(message)
        return

    if cur_user_id in users:
        data = "\n".join([f"{i[0]}: {i[1]}" for i in users[cur_user_id].items()])
        if cur_user_id == str(user_id):
            bot.send_message(user_id, f'Ваши текущие данные:\n{data}')
        else:
            bot.send_message(user_id, f'Текущие данные пользователя id{cur_user_id}:\n{data}')
        companies_list = [users[cur_user_id][COMPANY]]
        if user_id in ADMINS:
            companies_list += [i for i in companies if i != users[cur_user_id][COMPANY]]
    else:
        companies_list = []

    recording_data[user_id] = {USER_ID: cur_user_id}

    if user_id in ADMINS:
        bot.send_message(user_id, f'Введите название компании. Если компания еще не зарегистрирована, '
                                  f'вы можете добавить её с помощью команды /add_company',
                         reply_markup=make_keyboard(companies_list))
    else:
        bot.send_message(user_id, 'Введите название своей компании',
                         reply_markup=make_keyboard(companies_list))

    bot.register_next_step_handler(message, edit_user_company)


def edit_user_company(message):
    """Администратор редактирует данные зарегистрированного пользователя по его id"""
    if log(message):
        del recording_data[message.from_user.id]
        return

    user_id = message.from_user.id
    cur_data = recording_data[user_id]
    cur_user_id = cur_data[USER_ID]

    if message.text not in companies:
        bot.send_message(message.from_user.id, 'Компания с таким названием не зарегистрирована в системе. Обратитесь к администратору')
        bot.register_next_step_handler(message, edit_user_company)
        return

    cur_data[COMPANY] = message.text

    if cur_user_id in users and message.text == users[cur_user_id][COMPANY]:
        user = users[cur_user_id]
        addresses = [user[ADDRESS]] + [i for i in companies[user[COMPANY]] if i != user[ADDRESS]]
    else:
        addresses = companies[message.text]

    bot.send_message(user_id, f'Выберите адрес установки счетчика из списка или напишите новый',
                     reply_markup=make_keyboard(addresses))
    bot.register_next_step_handler(message, edit_user_address)


def edit_user_address(message):
    """Администратор редактирует данные зарегистрированного пользователя по его id"""
    if log(message):
        del recording_data[message.from_user.id]
        return

    user_id = message.from_user.id
    cur_data = recording_data[user_id]
    cur_user_id = cur_data[USER_ID]

    if message.text not in companies[cur_data[COMPANY]]:
        bot.send_message(user_id, 'Этот адрес будет внесен в список адресов компании')

    cur_data[ADDRESS] = message.text

    bot.send_message(user_id, f'Введите номер телефона в федеральном формате (+7**********)',
                     reply_markup=make_keyboard([users[cur_data[USER_ID]][PHONE]] if cur_user_id in users else []))
    bot.register_next_step_handler(message, edit_user_phone)


def edit_user_phone(message):
    """Администратор редактирует данные зарегистрированного пользователя по его id"""
    if log(message):
        del recording_data[message.from_user.id]
        return

    user_id = message.from_user.id
    cur_data = recording_data[user_id]
    cur_user_id = cur_data[USER_ID]

    try:
        cur_data[PHONE] = check_number(message.text)
    except (AssertionError, TypeError, ValueError):
        bot.send_message(user_id, 'Номер телефона не соответствует формату. Попробуйте ещё раз')
        bot.register_next_step_handler(message, edit_user_phone)
        return

    if cur_user_id in users:
        # Предлагаем имя, под которым он был зарегистрирован в прошлый раз (если был)
        names = [users[cur_user_id][USERNAME]]
    else:
        names = []

    if cur_user_id == str(user_id):
        # Предлагаем имя, под которым он зарегистрирован в Телеграмме
        fullname = ((message.from_user.last_name or ' ') + ' ' + message.from_user.first_name).strip()
        if not names or names[0] != fullname:
            names += [fullname]

    bot.send_message(message.from_user.id, 'Введите имя', reply_markup=make_keyboard(names))

    bot.register_next_step_handler(message, edit_user_username)


def edit_user_username(message):
    """Администратор редактирует данные зарегистрированного пользователя по его id"""
    if log(message):
        del recording_data[message.from_user.id]
        return

    user_id = message.from_user.id
    cur_data = recording_data[user_id]
    cur_user_id = cur_data[USER_ID]

    cur_data[USERNAME] = message.text

    if cur_user_id in users:
        changes = get_changes(users[cur_user_id], cur_data)
        bot.send_message(user_id, f'Вы подтверждаете изменения данных?\n{changes}',
                         reply_markup=make_bool_keyboard())
    else:
        changes = get_changes(cur_data)
        bot.send_message(user_id, f'Вы подтверждаете внесение данных?\n{changes}',
                         reply_markup=make_bool_keyboard())

    bot.register_next_step_handler(message, edit_user_verification)


def edit_user_verification(message):
    """Подтверждение внесения данных у пользователя"""
    if log(message):
        del recording_data[message.from_user.id]
        return

    user_id = message.from_user.id
    cur_data = recording_data[user_id]
    cur_user_id = cur_data[USER_ID]

    if message.text.lower() in POSITIVE_ANSWERS:
        if cur_data[ADDRESS] not in companies[cur_data[COMPANY]]:
            companies[cur_data[COMPANY]][cur_data[ADDRESS]] = {}
            dump(companies, COMPANIES_NAME)

        if cur_user_id in users:
            changes = get_changes(users[cur_user_id], cur_data)
        else:
            changes = ''

        # Перезапись файла
        del cur_data[USER_ID]
        users[cur_user_id] = cur_data.copy()
        dump(users, USERS_NAME)
        del recording_data[message.from_user.id]

        text1 = 'Если вы получили это, вы нашли баг (фичу) в программе'
        if changes and str(user_id) != cur_user_id:    # Администратор внес изменения
            bot.send_message(user_id, f'Вы успешно изменили данные пользователя id{cur_user_id}')
            text = f'Данные пользователя id{cur_user_id} были изменены администратором id{user_id}:\n{changes}'
            text1 = f'Ваши были изменены администратором:\n{changes}'

        elif changes:    # Сам пользователь внес изменения
            bot.send_message(user_id, 'Вы успешно изменили свои данные')
            text = f'Пользователь id{cur_user_id} изменил свои данные:\n{changes}'

        else:    # Пользователь зарегистрировался
            bot.send_message(user_id, 'Вы успешно зарегистрировались!')
            text = f'Пользователь id{cur_user_id} зарегистрировался:\n{get_changes(users[cur_user_id])}'

        # Отправляем информацию об изменениях админам и пользователю, данные которого были изменены
        for admin_id in set(ADMINS + [int(cur_user_id)]):
            if admin_id != user_id:
                bot.send_message(admin_id, text if admin_id != int(cur_user_id) else text1)

    else:
        del recording_data[message.from_user.id]
        bot.send_message(user_id, 'Возврат в меню')
    print_commands(message)


def remove_user_by_id(message):
    """Администратор удаляет зарегистрированного пользователя по его id"""
    if log(message):
        return

    if message.text.lower() == 'self':
        message.text = str(message.from_user.id)

    if message.text not in users:
        bot.send_message(message.from_user.id, 'Пользователь с этим id не зарегистрирован')
        print_commands(message)
        return

    recording_data[message.from_user.id] = message.text

    data = "\n".join([": ".join(map(str, i)) for i in [('id', message.text)] + list(users[message.text].items())])
    bot.send_message(message.from_user.id, f'Вы действительно хотите удалить пользователя с данными:\n{data}',
                     reply_markup=make_bool_keyboard())
    bot.register_next_step_handler(message, remove_user_verification)


def remove_user_verification(message):
    """Подтверждение удаления пользователя"""
    if log(message):
        del recording_data[message.from_user.id]
        return

    if message.text.lower() in POSITIVE_ANSWERS:
        del_user_id = recording_data[message.from_user.id]

        del users[del_user_id]
        dump(users, USERS_NAME)

        bot.send_message(int(del_user_id), 'Вы были удалены администратором.')

        for admin_id in ADMINS:
            if admin_id == message.from_user.id:
                text = f'Пользователь id{del_user_id} был удален.'
            else:
                text = f'Пользователь id{del_user_id} был удален администратором id{message.from_user.id}.'
            bot.send_message(admin_id, text)

        del recording_data[message.from_user.id]

    else:
        del recording_data[message.from_user.id]
        bot.send_message(message.from_user.id, 'Хорошо')


def add_company(message):
    """Регистрация компании администратором"""
    if log(message):
        return

    recording_data[message.from_user.id] = message.text

    if message.text in companies:
        bot.send_message(message.from_user.id, f'Компания "{message.text}" уже существует')
        print_commands(message)
        return

    bot.send_message(message.from_user.id, f'Зарегистрировать компанию "{message.text}"?',
                     reply_markup=make_bool_keyboard())
    bot.register_next_step_handler(message, add_company_verification)


def add_company_verification(message):
    """Подтверждение регистрации компании"""
    if log(message):
        del recording_data[message.from_user.id]
        return

    if message.text.lower() in POSITIVE_ANSWERS:
        companies[recording_data[message.from_user.id]] = {}
        dump(companies, COMPANIES_NAME)

        del recording_data[message.from_user.id]

        bot.send_message(message.from_user.id, 'Компания зарегистрирована')

    else:
        bot.send_message(message.from_user.id, 'Компания не зарегистрирована')
        message.text = 'exit_code_1'
        start(message)


def add_counter(message):
    """Регистрация прибора учета пользователем"""
    if log(message):
        return

    user = users[str(message.from_user.id)]
    if message.text in companies[user[COMPANY]][user[ADDRESS]]:
        bot.send_message(message.from_user.id, f'Прибор учета с номером "{message.text}" '
                                               f'уже зарегистрирован по адресу {user[ADDRESS]}')
        return

    recording_data[message.from_user.id] = message.text
    bot.send_message(message.from_user.id, f'Зарегистрировать прибор учета с номером "{message.text}" '
                                           f'по адресу "{users[str(message.from_user.id)][ADDRESS]}"?',
                     reply_markup=make_bool_keyboard())
    bot.register_next_step_handler(message, add_counter_verification)


def add_counter_verification(message):
    """Подтверждение регистрации прибора учета пользователем"""
    if log(message):
        del recording_data[message.from_user.id]
        return

    if message.text.lower() in POSITIVE_ANSWERS:
        user = users[str(message.from_user.id)]
        companies[user[COMPANY]][user[ADDRESS]][recording_data[message.from_user.id]] = ""
        dump(companies, COMPANIES_NAME)

        del recording_data[message.from_user.id]

        bot.send_message(message.from_user.id, 'Прибор учета зарегистрирован')

    else:
        bot.send_message(message.from_user.id, 'Прибор учета не зарегистрирован')
    print_commands(message)


def remove_counter(message):
    """Удаление прибора учета пользователем"""
    if log(message):
        return

    user_id = message.from_user.id
    user = users[str(user_id)]
    if message.text not in companies[user[COMPANY]][user[ADDRESS]]:
        bot.send_message(user_id, f'Прибор учета с номером "{message.text}" не зарегистрирован.')
        print_commands(message)
        return

    recording_data[user_id] = message.text
    bot.send_message(user_id, f'Удалить прибор учета с номером "{message.text}"?',
                     reply_markup=make_bool_keyboard())
    bot.register_next_step_handler(message, remove_counter_verification)


def remove_counter_verification(message):
    """Подтверждение удаления прибора учета пользователем"""
    if log(message):
        del recording_data[message.from_user.id]
        return

    if message.text.lower() in POSITIVE_ANSWERS:
        user = users[str(message.from_user.id)]
        del companies[user[COMPANY]][user[ADDRESS]][recording_data[message.from_user.id]]
        dump(companies, COMPANIES_NAME)

        del recording_data[message.from_user.id]

        bot.send_message(message.from_user.id, 'Прибор учета удален')

    else:
        bot.send_message(message.from_user.id, 'Прибор учета не удален')
    print_commands(message)


def get_counter(message):
    """Получение названия/номера счётчика у пользователя"""
    if log(message):
        return

    counter = message.text
    user = users[str(message.from_user.id)]

    if counter in companies[user[COMPANY]][user[ADDRESS]]:
        value = companies[user[COMPANY]][user[ADDRESS]][counter]
        if value:
            bot.send_message(message.from_user.id, f'Прошлое показание прибора учёта "{counter}": {value}.')
        else:
            bot.send_message(message.from_user.id, f'Нет предыдущих показаний по счётчику "{counter}"')

    else:
        bot.send_message(message.from_user.id, 'Прибор учёта с таким номером не зарегистрирован. '
                                               'Для регистрации введите /add_counter')
        return

    cur_data = recording_data[message.from_user.id] = {}
    cur_data[COMPANY] = user[COMPANY]
    cur_data[ADDRESS] = user[ADDRESS]
    cur_data[USERNAME] = user[USERNAME]
    cur_data[USER_ID] = str(message.from_user.id)
    cur_data[PHONE] = user[PHONE]
    cur_data[COUNTER] = counter

    bot.send_message(message.from_user.id, f'Введите текущее показание прибора учёта с номером "{counter}"')
    bot.register_next_step_handler(message, get_data)


def get_data(message):
    """Получение данных счётчика у пользователя"""
    if log(message):
        del recording_data[message.from_user.id]
        return

    if check_data(message.text):
        cur_data = recording_data[message.from_user.id]
        cur_data[DATA] = message.text
        cur_data[DATETIME] = get_date()

        s = "\n".join([": ".join(map(str, i)) for i in cur_data.items()])
        bot.send_message(message.from_user.id, f'Полученные данные: \n{s}')
        bot.send_message(message.from_user.id, f'Всё верно?', reply_markup=make_bool_keyboard())
        bot.register_next_step_handler(message, data_verification)

    else:
        bot.send_message(message.from_user.id, f'Введенные данные должны быть числом. Введите ещё раз')
        bot.register_next_step_handler(message, get_data)


def data_verification(message):
    """Подтверждение, что пользователь хочет внести данные"""
    if log(message):
        del recording_data[message.from_user.id]
        return

    if message.text.lower() not in POSITIVE_ANSWERS:
        del recording_data[message.from_user.id]
        bot.send_message(message.from_user.id, f'Данные не записаны')
        return

    cur_data = recording_data[message.from_user.id]

    companies[cur_data[COMPANY]][cur_data[ADDRESS]][cur_data[COUNTER]] = cur_data[DATA]
    dump(companies, COMPANIES_NAME)

    cursor = conn.cursor()
    # COMPANY, ADDRESS, USERNAME, USER_ID, PHONE, COUNTER, DATA, DATETIME = HEADERS
    cursor.execute(f"INSERT INTO records ({', '.join(HEADERS)}) VALUES (?, ?, ?, ?, ?, ?, ?, ?)",
                   [cur_data[header] for header in HEADERS])
    cursor.close()
    conn.commit()

    del recording_data[message.from_user.id]

    bot.send_message(message.from_user.id, f'Данные записаны')
    print_commands(message)


if __name__ == "__main__":
    while 1:
        try:
            bot.polling(none_stop=True, interval=0)
        except Exception as error:
            log_text = f'{get_date()} - FATAL_ERROR({error.__class__}, {error.__cause__}): {error}'
            print(log_text)
            open('log.txt', 'a', encoding='utf-8').write(log_text + '\n')
            sleep(1)
